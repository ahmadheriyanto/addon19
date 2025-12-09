from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from odoo.modules import get_module_path
import os
import os.path
from datetime import datetime
import base64
import xlrd
import math
import io
import re
import logging
_logger = logging.getLogger(__name__)

class ImportReceiptLine(models.TransientModel):
    _name = 'stock_picking_import_receipt'
    _description = 'import xls file into stock picking line'

    partner_id = fields.Many2one(
        'res.partner',
        'Partner',
        domain=['&', ('is_company', '=', True), ('category_id.name', 'ilike', 'Principal')]
    )
    location_id = fields.Many2one(
        'stock.location', "Source Location",
        default=lambda self: self.env['stock.picking.type'].browse(self.env.context.get('default_picking_type_id')).default_location_src_id, required=True)
    location_dest_id = fields.Many2one(
        'stock.location', "Destination Location",
        default=lambda self: self.env['stock.picking.type'].browse(self.env.context.get('default_picking_type_id')).default_location_dest_id, required=True)
    picking_type_id = fields.Many2one(
        'stock.picking.type', 'Operation Type', required=True)
    upload_file = fields.Binary(string="Lookup Excel File")
    fill_qty_done = fields.Boolean(string="Fill Quantity Done during import")

    @api.onchange('picking_type_id', 'partner_id')
    def onchange_picking_type(self):
        if self.picking_type_id:
            picking_type = self.env['stock.picking.type'].browse(self.picking_type_id.id)
            self.location_id = picking_type.default_location_src_id
            self.location_dest_id = picking_type.default_location_dest_id

        if self.partner_id and self.partner_id.picking_warn_msg:
            partner = self.partner_id
            return {'warning': {
                'title': ("Warning for %s") % partner.name,
                'message': partner.picking_warn_msg
            }}

    # It happens because the reserved quantity in your inventory does not reflect the one on your pickings.
    # It's probably due to a small configuration change while some pickings where open.
    # An easy way to remove it is to create a server action that executes this code
    def fix_unreserved_qty(self):
        quants = self.env["stock.quant"].search([])
        move_line_ids = []
        warning = ""

        for quant in quants:
            move_lines = self.env["stock.move.line"].search(
                [
                    ("product_id", "=", quant.product_id.id),
                    ("location_id", "=", quant.location_id.id),
                    ("lot_id", "=", quant.lot_id.id),
                    ("package_id", "=", quant.package_id.id),
                    ("owner_id", "=", quant.owner_id.id),
                    ("product_qty", "!=", 0),
                ]
            )

            move_line_ids += move_lines.ids
            reserved_on_move_lines = sum(move_lines.mapped("product_qty"))

            move_line_str = str.join(
                ", ", [str(move_line_id) for move_line_id in move_lines.ids]
            )

            if quant.location_id.should_bypass_reservation():
                # If a quant is in a location that should bypass the reservation, its `reserved_quantity` field
                # should be 0.
                if quant.reserved_quantity != 0:
                    quant.write({"reserved_quantity": 0})
            else:
                # If a quant is in a reservable location, its `reserved_quantity` should be exactly the sum
                # of the `product_qty` of all the partially_available / assigned move lines with the same
                # characteristics.

                if quant.reserved_quantity == 0:
                    if move_lines:
                        move_lines.with_context(bypass_reservation_update=True).write(
                            {"product_uom_qty": 0}
                        )
                elif quant.reserved_quantity < 0:
                    quant.write({"reserved_quantity": 0})
                    if move_lines:
                        move_lines.with_context(bypass_reservation_update=True).write(
                            {"product_uom_qty": 0}
                        )
                else:
                    if reserved_on_move_lines != quant.reserved_quantity:
                        move_lines.with_context(bypass_reservation_update=True).write(
                            {"product_uom_qty": 0}
                        )
                        quant.write({"reserved_quantity": 0})
                    else:
                        if any(move_line.product_qty < 0 for move_line in move_lines):
                            move_lines.with_context(bypass_reservation_update=True).write(
                                {"product_uom_qty": 0}
                            )
                            quant.write({"reserved_quantity": 0})

        move_lines = self.env["stock.move.line"].search(
            [
                ("product_id.type", "=", "product"),
                ("product_qty", "!=", 0),
                ("id", "not in", move_line_ids),
            ]
        )

        move_lines_to_unreserve = []

        for move_line in move_lines:
            if not move_line.location_id.should_bypass_reservation():
                move_lines_to_unreserve.append(move_line.id)

        if len(move_lines_to_unreserve) > 1:
            self.env.cr.execute(

                """ 

                    UPDATE stock_move_line SET product_uom_qty = 0, product_qty = 0 WHERE id in %s ;

                """

                % (tuple(move_lines_to_unreserve),)
            )

        elif len(move_lines_to_unreserve) == 1:
            self.env.cr.execute(

                """ 

                UPDATE stock_move_line SET product_uom_qty = 0, product_qty = 0 WHERE id = %s ;

                """

                % (move_lines_to_unreserve[0])

            )

    # Routine ini selalu create new document, bukan konsumsi untuk di panggil dari document, routine ini dipanggil dari menu action
    def import_format_inbound_wh(self):
        if not self.upload_file:
            raise UserError(_('Lookup xls excel file before upload'))

        if not (self.picking_type_id.code in ('incoming', 'internal', 'outgoing')):
            raise UserError(_('Operation Type is not inbound scope'))

        mpath = get_module_path('fulfillment')
        
        now = datetime.now()
        dt_string = now.strftime("%Y_%m_%d_%H_%M_%S_%f")
        out_file_name = _('inbound_%s.xls' % dt_string)
        
        out_file = mpath + _('/tmp/' + out_file_name)
        # delete file if exist
        if os.path.exists(out_file):
            os.remove(out_file)
        data = base64.b64decode(self.upload_file)
        with open(out_file, 'wb') as file:
            file.write(data)
        xl_workbook = xlrd.open_workbook(file.name)
        sheet_names = xl_workbook.sheet_names()
        sheetname = 'Sheet1'
        if not (sheetname in sheet_names):
            raise UserError(
                _('Worksheet with name "%s" does not exist' % sheetname))
        xl_sheet = xl_workbook.sheet_by_name(sheetname)
        # Number of Columns
        num_cols = xl_sheet.ncols
        # header
        headers = []
        for col_idx in range(0, num_cols):
            cell_obj = xl_sheet.cell(0, col_idx)
            headers.append(cell_obj.value)
        import_data = []
        for row_idx in range(1, xl_sheet.nrows):
            row_dict = {}
            for col_idx in range(0, num_cols):
                cell_obj = xl_sheet.cell(row_idx, col_idx)
                row_dict[headers[col_idx]] = cell_obj.value
            import_data.append(row_dict)

        stock_picking_id = []
        stock_picking = []
        if import_data:    
            # Prepare context for create (so default_get can pick defaults if used)
            ctx = dict(self.sudo().env.context or {})
            # Prevent stock.move merge so each incoming line stays separate
            ctx['no_merge'] = True
            for row in import_data:
                partner_type = row['PARTNER TYPE (B2B/B2C)'].lower()
                # Check master product
                product_no = row['PRODUCT CODE']
                if isinstance(product_no, int) or isinstance(product_no, float):
                    fractional, whole = math.modf(product_no)
                    if fractional == 0:
                        product_no = str(int(product_no))
                    else:
                        product_no = str(product_no)

                check_product = self.env['product.product'].search(
                    [('default_code', '=', product_no)])
                if not check_product:
                    raise UserError(
                        _('Product %s does not exist in master data' % product_no))
                else:
                    check_product = check_product[0]
                    
                #raise UserError(_('System Sedang di supervisi oleh Ahmad: sukses check_product '))
            
                # Check master UoM
                check_uom = self.env['uom.uom'].search(
                    [('name', '=', row['UOM'])])
                if not check_uom:
                    raise UserError(
                        _('UoM %s does not exist in master data' % row['UOM']))
                else:
                    check_uom = check_uom[0]
                #                   0                   1                2               3             4             5                    6
                # rowdata = [check_product.id,row['PRODUCT CODE'],row['PO NUMBER'],check_uom.id,row['UOM'],row['BATCH'],row['EXPIRED'],row['QTY']
                ijno = row['PO NUMBER']
                stock_picking = self.env['stock.picking'].search(
                    [('origin', '=', ijno), ('picking_type_id', '=', self.picking_type_id.id)])
                if not stock_picking:
                    stock_picking = stock_picking.create({
                        'partner_id': self.partner_id.id,
                        'picking_type_id': self.picking_type_id.id,
                        'location_id': self.location_id.id,
                        'location_dest_id': self.location_dest_id.id,
                        'origin': ijno,
                        'partner_type': partner_type,
                    })
                else:
                    stock_picking = stock_picking[0]
                    if stock_picking.state in ['done', 'cancel']:
                        raise UserError(_('Stock Picking with Source Document %s already exist in the system with status %s, Source Document must be on other new value' % (
                            ijno, stock_picking.state)))

                if not (stock_picking.id in stock_picking_id):
                    stock_picking_id.append(stock_picking.id)

                # special for putaway
                loc_obj = []
                if self.picking_type_id.code == 'internal':                    
                    loc_obj = self.env['stock.location'].search([
                        ('complete_name', '=', _('%s' %
                            row['DESTINATION LOC'])),
                    ])
                    if not loc_obj:
                        raise UserError(
                            _('Location %s is not found in database' % row['DESTINATION LOC']))

                # Check Stock.lot
                #raise UserError(_('test: ijno = %s, Best before date = %s ' % (ijno,datetime(*xlrd.xldate_as_tuple(row['Best before date'], 0)))))
                batchno = row['BATCH']
                if isinstance(batchno, int) or isinstance(batchno, float):
                    fractional, whole = math.modf(batchno)
                    if fractional == 0:
                        batchno = str(int(batchno))
                    else:
                        batchno = str(batchno)

                check_lot = self.env['stock.lot'].search(
                    [('name', '=', batchno), ('product_id', '=', check_product.id)])
                if not check_lot:
                    check_lot = self.env['stock.lot'].create({
                        'name': batchno,
                        'product_id': check_product.id,
                        'ref': ijno,
                        'use_expiration_date': True,
                        'expiration_date': datetime(*xlrd.xldate_as_tuple(row['EXPIRED'], 0))
                    })
                else:
                    check_lot = check_lot[0]

                qty_done = row['QTY']

                # mulai input data
                stock_move = stock_picking.move_ids.with_context(ctx).create({
                    'description_picking': f"{check_product.display_name} [LOT: {check_lot.name}]",
                    'sequence': 10,
                    'company_id': stock_picking.company_id.id,
                    'product_id': check_product.id,
                    'product_uom': check_uom.id,
                    # 'product_qty': row['QTY Kirim'],
                    'product_uom_qty': row['QTY'],
                    'quantity': qty_done,
                    'location_id': stock_picking.location_id.id,
                    'location_dest_id': (
                        loc_obj[0].id if (self.picking_type_id.code in ('incoming', 'internal') and loc_obj)
                        else stock_picking.location_dest_id.id
                    ),
                    'picking_type_id': stock_picking.picking_type_id.id,
                    'picking_id': stock_picking.id,
                    #'lot_ids': [(6, 0, check_lot.ids)],
                })   

                if check_lot:
                    stock_move.sudo().write({'lot_ids': [(6, 0, check_lot.ids)]})             

            #raise UserError(_('Sukses import_data, stock_picking_id = %s' % stock_picking_id))
            if (self.picking_type_id.code == 'incoming'):
                #self.fix_unreserved_qty()  # supaya tidak terjadi reserve
                for picking_id in stock_picking_id:
                    pick = self.env["stock.picking"].search([('id', '=', picking_id)])
                    pick.action_confirm() #mark as todo
                    pick.action_assign()  #check avaibility
                    
                    # # *** Manage QR Data ***
                    # qr_data = self.env['qr.data.header'].search([('picking_id', '=', pick.id)])
                    # if not qr_data:
                    #     qr_data = qr_data.create({
                    #         'name': pick.name,
                    #         'picking_type_id': pick.picking_type_id.id,
                    #         'description': pick.origin,
                    #         'picking_id':pick.id,
                    #     })
                    # for stock_move_for_qrdata in pick.move_ids:
                    #     qr_data_line = self.env['qr.data.line'].search([('picking_move_id', '=', stock_move_for_qrdata.id)])
                    #     if not qr_data_line:
                    #         qr_data.line_ids.create({
                    #             'header_id': qr_data.id,
                    #             'picking_move_id': stock_move_for_qrdata.id,
                    #             'product_id': stock_move_for_qrdata.product_id.id,
                    #             'lot_name': stock_move_for_qrdata.suggest_lot_id.name,
                    #             'expired_date': stock_move_for_qrdata.suggest_lot_id.use_date,
                    #             'product_qty': stock_move_for_qrdata.product_uom_qty,
                    #             'product_uom': stock_move_for_qrdata.product_uom.id,
                    #         })
                    # qr_data.action_generate_qr()
        
            #Show result in list based on stock_picking_id
            context = dict(self.env.context)
            #context['form_view_initial_mode'] = 'edit'
            return {
                'domain': "[('id','in', ["+','.join(map(str, stock_picking_id))+"])]", # stock_picking.ids
                'name': _('New Created Records (Inbound)'),
                'view_type': 'form',
                'view_mode': 'list,form',
                'res_model': 'stock.picking',
                'view_id': False,
                'context': False,
                'type': 'ir.actions.act_window'
            }

    # Routine ini selalu create new document, bukan konsumsi untuk di panggil dari document, routine ini dipanggil dari menu action
    def import_format_outbound_wh(self):
        if not self.upload_file:
            raise UserError(_('Lookup xls excel file before upload'))
        mpath = get_module_path('fulfillment')
        
        # out_file_name = 'inbound.xls'
        now = datetime.now()
        dt_string = now.strftime("%Y_%m_%d_%H_%M_%S_%f")
        out_file_name = _('outbound_%s.xls' % dt_string)
        
        out_file = mpath + _('/tmp/' + out_file_name)
        # delete file if exist
        if os.path.exists(out_file):
            os.remove(out_file)
        data = base64.b64decode(self.upload_file)
        with open(out_file, 'wb') as file:
            file.write(data)
        xl_workbook = xlrd.open_workbook(file.name)
        sheet_names = xl_workbook.sheet_names()
        sheetname = 'Sheet1'
        if not (sheetname in sheet_names):
            raise UserError(
                _('Worksheet with name "%s" does not exist' % sheetname))
        xl_sheet = xl_workbook.sheet_by_name(sheetname)
        # Number of Columns
        num_cols = xl_sheet.ncols
        # header
        headers = []
        for col_idx in range(0, num_cols):
            cell_obj = xl_sheet.cell(0, col_idx)
            headers.append(cell_obj.value)
        import_data = []
        for row_idx in range(1, xl_sheet.nrows):
            row_dict = {}
            for col_idx in range(0, num_cols):
                cell_obj = xl_sheet.cell(row_idx, col_idx)
                row_dict[headers[col_idx]] = cell_obj.value
            import_data.append(row_dict)

        # Kumpulkan per Customer
        customerpls = []
        stock_picking_list = []
        stock_picking = self.env['stock.picking']
        if import_data:
            for row in import_data:
                customerno = row['CUSTOMER']
                if customerno:
                    if customerno != '':
                        customerpls.sort()
                        if not (customerno in customerpls):
                            customerpls.append(customerno)

        stock_picking_id = []
        for custno in customerpls:
            #_logger.critical('****** >>>> C E K: <<<<< ***** | custno = %s' % (custno))
            check_cust = self.env['res.partner'].search([('ref', '=', custno)])
            if not check_cust:
                raise UserError(
                    _('1. Customer Internal Reference %s does not exist in master data' % custno))
            else:
                check_cust = check_cust[0]
            stock_picking = self.env['stock.picking'].create({
                'partner_id': check_cust.id,
                'picking_type_id': self.picking_type_id.id,
                'location_id': self.location_id.id,
                'location_dest_id': self.location_dest_id.id
                # 'origin': ijno -> source document tidak perlu karena multi PLS di letakkan di line
            })

            if not (stock_picking.id in stock_picking_id):
                stock_picking_id.append(stock_picking.id)
            
            stock_picking_list.append({
                'stock_picking_id': stock_picking.id,
                'customer_id': check_cust.id
            })

        if import_data and stock_picking_list:
            for row in import_data:
                packing_list_no = row['PICKING NUMBER']
                packing_list_city = row['City']

                # Check master product
                product_no = row['PRODUCT CODE']
                if isinstance(product_no, int) or isinstance(product_no, float):
                    fractional, whole = math.modf(product_no)
                    if fractional == 0:
                        product_no = str(int(product_no))
                    else:
                        product_no = str(product_no)

                check_product = self.env['product.product'].search(
                    [('default_code', '=', product_no)])
                if not check_product:
                    raise UserError(
                        _('Product %s does not exist in master data' % product_no))
                else:
                    check_product = check_product[0]

                # Check master UoM
                check_uom = self.env['uom.uom'].search(
                    [('name', '=', row['UOM'])])
                if not check_uom:
                    raise UserError(
                        _('UoM %s does not exist in master data' % row['UOM']))
                else:
                    check_uom = check_uom[0]

                # Check Customer
                custno = '' + row['CUSTOMER']
                check_cust = self.env['res.partner'].search(
                    [('ref', '=', custno)])
                if not check_cust:
                    raise UserError(
                        _('2. Customer Internal Reference %s does not exist in master data' % custno))
                else:
                    check_cust = check_cust[0]

                # <<*** Header sudah di create per customer di coding atasnya
                stock_picking_per_customer = next(
                    (elmt for elmt in stock_picking_list if elmt['customer_id'] == check_cust.id), None)

                stock_picking_find = stock_picking.search(
                    [('id', '=', stock_picking_per_customer["stock_picking_id"])])
                stock_picking_find = stock_picking_find[0]
                
                # Check LOT

                # Check Stock.lot
                #raise UserError(_('test: ijno = %s, Best before date = %s ' % (ijno,datetime(*xlrd.xldate_as_tuple(row['Best before date'], 0)))))
                batchno = row['Batch Number']
                if isinstance(batchno, int) or isinstance(batchno, float):
                    fractional, whole = math.modf(batchno)
                    if fractional == 0:
                        batchno = str(int(batchno))
                    else:
                        batchno = str(batchno)

                check_lot = self.env['stock.lot'].search(
                    [('name', '=', batchno), ('product_id', '=', check_product.id)])
                if not check_lot:
                    raise UserError(
                        _('Batch# %s on Product %s does not exist in master data' % (batchno, product_no)))
                else:
                    check_lot = check_lot[0]

                # mulai input data
                if stock_picking_find:
                    stock_move = stock_picking_find.move_lines.create({
                        'name': check_product.name,
                        'sequence': 10,
                        'company_id': stock_picking_find.company_id.id,
                        'product_id': check_product.id,
                        'product_uom': check_uom.id,
                        # 'product_qty': row['QTY Kirim'],
                        'product_uom_qty': row['QTY'],
                        # 'reserved_avaibility': 0,
                        'location_id': stock_picking_find.location_id.id,
                        'location_dest_id': stock_picking_find.location_dest_id.id,
                        'picking_type_id': stock_picking_find.picking_type_id.id,
                        'packing_list_no': packing_list_no,
                        # 'packing_list_city': packing_list_city,
                        'picking_id': stock_picking_find.id,
                        'lot_ids': [(6, 0, check_lot.ids)],
                    })
                    

                    # Manage packing_list_no into header [origin field]
                    if stock_picking_find.origin:
                        if stock_picking_find.origin.find(packing_list_no) == -1:
                            stock_picking_find.origin = stock_picking_find.origin + '|' + packing_list_no
                    else:
                        stock_picking_find.origin = packing_list_no
                    
            #Show result in list based on stock_picking_id
            context = dict(self.env.context)
            #context['form_view_initial_mode'] = 'edit'
            return {
                'domain': "[('id','in', ["+','.join(map(str, stock_picking_id))+"])]", # stock_picking.ids
                'name': _('New Created Records (Outbound)'),
                'view_type': 'form',
                'view_mode': 'tree,form',
                'res_model': 'stock.picking',
                'view_id': False,
                'context': False,
                'type': 'ir.actions.act_window'
            }

    # @api.model
    def export_template_importdata(self, inspection_id=None):
        """
        Build an XLSX template and, if possible, prefill a data row from:
         - provided inspection_id, or
         - active_id in context, or
         - first record in inspection_bpkb (ordered by `no` ascending).

        Returns an ir.actions.act_url to download the generated file.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise UserError(_("openpyxl must be installed on the server to export the template."))

        def sanitize_sheet_title(title: str) -> str:
            title = re.sub(r'[:\\\/\?\*\[\]]', '', (title or '').strip())
            if len(title) > 31:
                title = title[:31]
            if not title:
                title = "Sheet1"
            return title

        # Build workbook and headers (must match header_check mapping)
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sanitize_sheet_title("Template")
        
        headers = [
            'PO NUMBER', 'PRODUCT CODE', 'BATCH', 'EXPIRED',
            'QTY', 'UOM', 'PARTNER TYPE (B2B/B2C)'
        ]
        if self.picking_type_id.code == 'internal':
            headers = [
                'PO NUMBER', 'PRODUCT CODE', 'BATCH', 'EXPIRED',
                'QTY', 'UOM', 'DESTINATION LOC', 'PARTNER TYPE (B2B/B2C)'
            ]    
        ws.append(headers)

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        template_bytes = buf.getvalue()

        # Create a temporary attachment and return URL to download it
        attachment = self.env['ir.attachment'].create({
            'name': 'inbound.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(template_bytes).decode('utf-8'),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': False,
            'res_id': False,
            'public': False,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
